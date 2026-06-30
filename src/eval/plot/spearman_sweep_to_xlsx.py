#!/usr/bin/env python3
"""
Export Spearman sweep tables (dyn threshold × category) to Excel, one sheet per model.

Layout matches the template: merged model name on row 1, category headers on row 2,
thresholds 0.2–0.8 in column A, Spearman ρ in the grid.

Requires: click, openpyxl (stdlib csv only — no pandas).

Usage (from repo root):

  python src/eval/plot/spearman_sweep_to_xlsx.py \\
    data_eval/plots/dynamic_legacy/msa_0.3/spearman_table_msa_0.3_legacy.csv \\
    data_eval/plots/dynamic_legacy/msa_0.3/spearman_sweep_by_model.xlsx
"""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Any

import click
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

MODEL_LABELS: dict[str, str] = {
    "af3": "AF3",
    "boltz1": "Boltz-1",
    "boltz2": "Boltz-2",
}

# (scope suffix after "model::", column title)
CATEGORIES: list[tuple[str, str]] = [
    ("all", "All"),
    ("intrinsic", "Intrinsic"),
    ("ligand-induced::apo", "Ligand-induced(Apo)"),
    ("ligand-induced::holo", "Ligand-induced(Holo)"),
    ("protein-induced::apo", "Protein-induced(Apo)"),
    ("protein-induced::holo", "Protein-induced(Holo)"),
]

THRESHOLDS: list[float] = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]


def load_spearman_grid(
    csv_path: Path,
) -> dict[str, dict[str, dict[float, float]]]:
    """model -> category -> dyn_threshold -> spearman_rho."""
    grid: dict[str, dict[str, dict[float, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    with csv_path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            scope = (row.get("scope") or "").strip()
            if "::" not in scope:
                continue
            model, cat = scope.split("::", 1)
            thr = float(row["dyn_threshold"])
            grid[model][cat][thr] = float(row["spearman_rho"])
    return {m: dict(cats) for m, cats in grid.items()}


def write_xlsx(
    grid: dict[str, dict[str, dict[float, float]]],
    path: Path,
) -> None:
    wb = Workbook()
    first = True
    cat_keys = [c[0] for c in CATEGORIES]
    col_labels = [c[1] for c in CATEGORIES]

    for model in sorted(grid.keys()):
        sheet_name = MODEL_LABELS.get(model, model)[:31]
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        ncols = len(col_labels)
        ws.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=1 + ncols,
        )
        c = ws.cell(row=1, column=2, value=sheet_name)
        c.alignment = Alignment(horizontal="center", vertical="center")

        for j, name in enumerate(col_labels, start=2):
            ws.cell(row=2, column=j, value=name)

        mg = grid[model]
        for i, thr in enumerate(THRESHOLDS):
            r = 3 + i
            ws.cell(row=r, column=1, value=thr)
            for j, ck in enumerate(cat_keys, start=2):
                val: Any = mg.get(ck, {}).get(thr)
                if val is None:
                    ws.cell(row=r, column=j, value="")
                else:
                    ws.cell(row=r, column=j, value=float(val))

        ws.column_dimensions["A"].width = 12
        for j in range(2, 2 + ncols):
            ws.column_dimensions[get_column_letter(j)].width = 22

    wb.save(path)


@click.command()
@click.argument(
    "csv_path",
    type=click.Path(exists=True, path_type=Path),
)
@click.argument(
    "xlsx_path",
    type=click.Path(path_type=Path),
)
def main(csv_path: Path, xlsx_path: Path) -> None:
    grid = load_spearman_grid(csv_path)
    if not grid:
        raise click.ClickException("No model-scoped rows found in CSV.")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(grid, xlsx_path)
    click.echo(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
