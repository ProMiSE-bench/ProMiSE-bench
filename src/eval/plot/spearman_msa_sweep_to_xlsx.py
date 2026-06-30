#!/usr/bin/env python3
"""
Aggregate Spearman ρ across MSA-pref threshold folders at a fixed DynDisto threshold.

One sheet per model (category columns). Rows sweep MSA pref thresholds (from msa_*/ folders).
With --show-n (default), each MSA uses two rows: Spearman ρ on the first row and N below.
If CSVs lack an `n` column, N cells show "-" until you regenerate tables from
plot_correlation_dynamic_legacy.py.

Requires: click, openpyxl (stdlib csv only — no pandas).

Usage (from repo root):

  python src/eval/plot/spearman_msa_sweep_to_xlsx.py \\
    data_eval/plots/dynamic_legacy \\
    data_eval/plots/dynamic_legacy/spearman_msa_sweep_dyn_0.7.xlsx \\
    --dyn-threshold 0.7
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import click
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

MODEL_LABELS: dict[str, str] = {
    "af3": "AF3",
    "boltz1": "Boltz-1",
    "boltz2": "Boltz-2",
}

CATEGORIES: list[tuple[str, str]] = [
    ("all", "All"),
    ("intrinsic", "Intrinsic"),
    ("ligand-induced::apo", "Ligand-induced(Apo)"),
    ("ligand-induced::holo", "Ligand-induced(Holo)"),
    ("protein-induced::apo", "Protein-induced(Apo)"),
    ("protein-induced::holo", "Protein-induced(Holo)"),
]

MSA_DIR_RE = re.compile(r"^msa_(.+)$")


def _parse_msa_from_dir(name: str) -> float | None:
    m = MSA_DIR_RE.match(name)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def peek_any_csv_has_n_column(plot_dir: Path) -> bool:
    """True if at least one spearman_table CSV has an `n` column in the header."""
    msa_dirs = [
        p for p in plot_dir.iterdir()
        if p.is_dir() and _parse_msa_from_dir(p.name) is not None
    ]
    msa_dirs.sort(key=lambda p: _parse_msa_from_dir(p.name) or 0.0)
    for msa_dir in msa_dirs:
        p = _find_spearman_csv(msa_dir)
        if p is None:
            continue
        with p.open(newline="", encoding="utf-8") as f:
            header = next(csv.reader(f))
        return "n" in header
    return False


def _find_spearman_csv(msa_dir: Path) -> Path | None:
    """Prefer spearman_table_msa_<T>_legacy.csv matching folder name."""
    t = _parse_msa_from_dir(msa_dir.name)
    if t is None:
        return None
    exact = msa_dir / f"spearman_table_msa_{t}_legacy.csv"
    if exact.is_file():
        return exact
    cands = sorted(msa_dir.glob("spearman_table_*_legacy.csv"))
    return cands[0] if cands else None


def load_slice_at_dyn(
    csv_path: Path,
    dyn_fixed: float,
) -> dict[str, dict[str, tuple[float, int | None]]]:
    """model -> category -> (spearman_rho, n) at dyn_threshold == dyn_fixed."""
    out: dict[str, dict[str, tuple[float, int | None]]] = {}
    with csv_path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            scope = (row.get("scope") or "").strip()
            if "::" not in scope:
                continue
            if abs(float(row["dyn_threshold"]) - dyn_fixed) > 1e-9:
                continue
            model, cat = scope.split("::", 1)
            if model not in out:
                out[model] = {}
            n_raw = (row.get("n") or "").strip()
            n_val = int(n_raw) if n_raw else None
            out[model][cat] = (float(row["spearman_rho"]), n_val)
    return out


def collect_msa_sweep(
    plot_dir: Path,
    dyn_fixed: float,
) -> tuple[list[float], dict[str, dict[str, dict[float, tuple[float, int | None]]]]]:
    """
    Returns (sorted msa thresholds, grid model -> category -> msa_thr -> rho).
    """
    msa_dirs = [
        p for p in plot_dir.iterdir()
        if p.is_dir() and _parse_msa_from_dir(p.name) is not None
    ]
    msa_dirs.sort(key=lambda p: _parse_msa_from_dir(p.name) or 0.0)
    msa_thresholds: list[float] = []
    grid: dict[str, dict[str, dict[float, tuple[float, int | None]]]] = {}

    for msa_dir in msa_dirs:
        msa_t = _parse_msa_from_dir(msa_dir.name)
        if msa_t is None:
            continue
        csv_path = _find_spearman_csv(msa_dir)
        if csv_path is None:
            continue
        slice_ = load_slice_at_dyn(csv_path, dyn_fixed)
        if not slice_:
            continue
        msa_thresholds.append(msa_t)
        for model, cats in slice_.items():
            if model not in grid:
                grid[model] = {}
            for cat, val in cats.items():
                if cat not in grid[model]:
                    grid[model][cat] = {}
                grid[model][cat][msa_t] = val

    msa_thresholds = sorted(set(msa_thresholds))
    return msa_thresholds, grid


def write_xlsx(
    msa_thresholds: list[float],
    grid: dict[str, dict[str, dict[float, tuple[float, int | None]]]],
    path: Path,
    show_n: bool,
) -> None:
    wb = Workbook()
    first = True
    cat_keys = [c[0] for c in CATEGORIES]
    col_labels = [c[1] for c in CATEGORIES]

    for model in sorted(grid.keys()):
        sheet_name = MODEL_LABELS.get(model, model)[:31]
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        ncols = len(col_labels)
        ws.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=1 + ncols,
        )
        c = ws.cell(row=1, column=2, value=sheet_name)
        c.alignment = Alignment(horizontal="center", vertical="center")

        mg = grid[model]
        if show_n:
            # Rows 2–3: merged headers (room for ρ / N sub-rows). Data starts row 4.
            ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
            msa_hdr = ws.cell(row=2, column=1, value="MSA")
            msa_hdr.alignment = Alignment(horizontal="center", vertical="center")
            for j in range(2, 2 + ncols):
                ws.merge_cells(start_row=2, start_column=j, end_row=3, end_column=j)
                hdr = ws.cell(row=2, column=j, value=col_labels[j - 2])
                hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_start = 4
            for i, msa_t in enumerate(msa_thresholds):
                r_rho = data_start + 2 * i
                r_n = r_rho + 1
                ws.merge_cells(
                    start_row=r_rho,
                    start_column=1,
                    end_row=r_n,
                    end_column=1,
                )
                a = ws.cell(row=r_rho, column=1, value=msa_t)
                a.alignment = Alignment(vertical="center")
                for j, ck in enumerate(cat_keys, start=2):
                    val = mg.get(ck, {}).get(msa_t)
                    if val is None:
                        ws.cell(row=r_rho, column=j, value="")
                        ws.cell(row=r_n, column=j, value="")
                    else:
                        rho, n = val
                        ws.cell(row=r_rho, column=j, value=float(rho))
                        if n is not None:
                            ws.cell(row=r_n, column=j, value=int(n))
                        else:
                            ws.cell(row=r_n, column=j, value="-")
        else:
            for j, name in enumerate(col_labels, start=2):
                ws.cell(row=2, column=j, value=name)
            for i, msa_t in enumerate(msa_thresholds):
                r = 3 + i
                ws.cell(row=r, column=1, value=msa_t)
                for j, ck in enumerate(cat_keys, start=2):
                    val: Any = mg.get(ck, {}).get(msa_t)
                    if val is None:
                        ws.cell(row=r, column=j, value="")
                    else:
                        rho, _n = val
                        ws.cell(row=r, column=j, value=float(rho))

        ws.column_dimensions["A"].width = 14
        for j in range(2, 2 + ncols):
            ws.column_dimensions[get_column_letter(j)].width = 22

    wb.save(path)


@click.command()
@click.argument(
    "plot_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
)
@click.argument("xlsx_path", type=click.Path(path_type=Path))
@click.option(
    "--dyn-threshold",
    type=float,
    default=0.7,
    show_default=True,
    help="Fixed DynDisto threshold; one row is taken per MSA folder at this value.",
)
@click.option(
    "--show-n/--hide-n",
    default=True,
    show_default=True,
    help="Second row per MSA for N; if CSV has no 'n' column, N cells are '-'.",
)
def main(plot_dir: Path, xlsx_path: Path, dyn_threshold: float, show_n: bool) -> None:
    msa_thresholds, grid = collect_msa_sweep(plot_dir, dyn_threshold)
    if not msa_thresholds or not grid:
        raise click.ClickException(
            "No data: need msa_*/ folders with spearman_table_*_legacy.csv "
            f"containing dyn_threshold={dyn_threshold}."
        )
    if show_n and not peek_any_csv_has_n_column(plot_dir):
        click.echo(
            "Warning: no spearman_table CSV under this plot_dir has an 'n' column. "
            "N rows will show '-'. Regenerate CSVs with plot_correlation_dynamic_legacy.py "
            "(--table-output-csv).",
            err=True,
        )
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(msa_thresholds, grid, xlsx_path, show_n)
    click.echo(
        f"Wrote {xlsx_path} (MSA thresholds: {msa_thresholds}, dyn fixed: {dyn_threshold}, show_n: {show_n})"
    )


if __name__ == "__main__":
    main()
